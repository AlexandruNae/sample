# import xlsxwriter
import json
import datetime
from datetime import date, timedelta
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

now = datetime.datetime.now()
folder_an = now.year
folder_luna = now.month
folder_zi = now.day

dict_numar_nume_luna = {
  "1": "ianuarie",
  "2": "februarie",
  "3": "martie",
  "4": "aprilie",
  "5": "mai",
  "6": "iunie",
  "7": "iulie",
  "8": "august",
  "9": "septembrie",
  "10": "octombrie",
  "11": "noiembrie",
  "12": "decembrie"
}

nume_luna = dict_numar_nume_luna[str(folder_luna)]
folder = str(folder_an) + "\\" + nume_luna + "\\" + str(folder_zi)
# folder = str(folder_an) + "\\" + nume_luna + "\\" + 'test4'

directory_error = 'C:\\Users\\quasar\\invoice_handling\\'+ folder +'\\5_VALIDATE_INVOICE\\'
xlsx_file = 'C:\\Users\\quasar\\invoice_handling\\'+ folder +'\\'+ dict_numar_nume_luna[str(folder_luna)] + '_' + str(folder_zi) + '_pass' +'.xlsx'

count=1
for file in os.listdir(directory_error):
    with open(directory_error+file) as json_file:
        raw_data = json.load(json_file)

    data = {
        'nume_factura':raw_data['nume_factura'],
        'data':raw_data['data'],
        'numar_factura':raw_data['numar_factura'],
        'denumire_firma':raw_data['denumire_firma'],
        'denumire_medic':raw_data['denumire_medic'],
        'total':raw_data['total'],
        'grup':raw_data['grup'],
        'reason':raw_data['reason'],
        'reprezentant_hr':raw_data['reprezentant_hr']
    }
    try:
        wb = load_workbook(xlsx_file)
        ws = wb.worksheets[0]  # select first worksheet
    except FileNotFoundError:
        headers_row = ['nume_factura', 'data', 'numar_factura', 'denumire_firma', 'denumire_medic', 'total', 'grup', 'reason', 'reprezentant_hr']
        
        wb = Workbook()
        # wb = load_workbook(xlsx_file)
        ws = wb.active
        ws.append(headers_row)
        # Select First Worksheet
        

    row_data=[]
    for i, (k, v) in enumerate(data.items(), start=0):
        row_data.append(str(v))
    try:

        ws.append(row_data)
    except openpyxl.utils.exceptions.IllegalCharacterError as e:
        print(e)
    wb.save(xlsx_file)
    # with xlsxwriter.Workbook(xlsx_file) as workbook:

    #     # Add worksheet
    #     worksheet = workbook.add_worksheet()

    #     # Write headers
    #     worksheet.write(0, 0, 'nume_factura')
    #     worksheet.write(0, 1, 'data')
    #     worksheet.write(0, 2, 'numar_factura')
    #     worksheet.write(0, 3, 'denumire_firma')
    #     worksheet.write(0, 4, 'denumire_medic')
    #     worksheet.write(0, 5, 'total')
    #     worksheet.write(0, 6, 'grup')
    #     worksheet.write(0, 7, 'reason')
    #     worksheet.write(0, 8, 'reprezentant_hr')

    #     for i, (k, v) in enumerate(data.items(), start=0):
    #         worksheet.write(count, i, str(v))
    # count +=1
